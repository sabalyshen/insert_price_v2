import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
# loadfile
wb = load_workbook(filename = 'sheet.xlsx', data_only=False)
ws = wb['大隊日記帳']

wb_t = load_workbook(filename = 'work_sample.xlsx', data_only=True)
ws_t = wb_t['sheet']

wb_a = load_workbook(filename = '110年零用金流向0610.xlsx', data_only=True)
ws_a = wb_a['work']

colA = ws_a['A']
list_colA = []
for c in colA:
    if c.value != None:
        list_colA.append(c)
print(list_colA)
print(len(list_colA))

#border
bian = Side(style='thin', color='000000') 
bian1 = Side(style='medium', color='000000')
border = Border(top=bian, bottom=bian, left=bian, right=bian, diagonal=bian, diagonalDown=True)
border1 = Border(top=bian, bottom=bian1, left=bian, right=bian, diagonal=bian, diagonalDown=True)
border3 = Border(top=bian, bottom=bian, left=bian, right=bian)

#準備投入的資料
type_name = []
with open('type.txt', 'r', encoding = 'utf8') as f:
    for n in f:
        n = n.replace('\n', '')
        type_name.append(n)
        print(n)
def wordfinder_sort(searchString):
    for cA in list_colA:
        if cA.value == searchString:
            if ws_a.cell(cA.row, 7).value == '物品':
                ws_f['U1'].value = '335-'
                ws_f['C6'].value = '消防業務－各救災救護大隊－業務費－物品'
                ws_f['P6'].value = ws_a.cell(cA.row, 8).value + '\n代墊人: ' 
            elif ws_a.cell(cA.row, 7).value == '一般':
                ws_f['U1'].value = '396-'
                ws_f['C6'].value = '消防業務－各救災救護大隊－業務費－一般事務費'
                ws_f['P6'].value = ws_a.cell(cA.row, 8).value + '\n代墊人: ' 
            elif ws_a.cell(cA.row, 7).value == '電費':
                ws_f['U1'].value = '213-'
                ws_f['C6'].value = '消防業務－各救災救護大隊－業務費－水電費'
                ws_f['P6'].value = ws_a.cell(cA.row, 8).value + '\n代墊人: ' 
            elif ws_a.cell(cA.row, 7).value == '水費':
                ws_f['U1'].value = '158-'
                ws_f['C6'].value = '消防業務－各救災救護大隊－業務費－水費'
                ws_f['P6'].value = ws_a.cell(cA.row, 8).value + '\n代墊人: ' 
            elif ws_a.cell(cA.row, 7).value == '電話':
                ws_f['U1'].value = '278-'
                ws_f['C6'].value = '消防業務－各救災救護大隊－業務費－通訊費'
                ws_f['P6'].value = ws_a.cell(cA.row, 8).value + '\n代墊人: '

def wordfinder_name(searchString):
    n = 0
    for cA in list_colA:
        if cA.value == searchString:
            ws_f.cell(21+n, 1).value= ws_a.cell(cA.row, 3).value
            n += 1 
          
def wordfinder_type(searchString):
    n = 0
    for cA in list_colA:
        if cA.value == searchString:
            ws_f.cell(21+n, 7).value= ws_a.cell(cA.row, 4).value
            n += 1 

def wordfinder_quantity(searchString):
    n = 0
    for cA in list_colA:
        if cA.value == searchString:
            ws_f.cell(21+n, 10).value= ws_a.cell(cA.row, 5).value
            n += 1 

def wordfinder_price(searchString):
    a = ws_f['A21'].value
    n = 0
    for cA in list_colA:
        if cA.value == searchString:
            ws_f.cell(21+n, 13).value= ws_a.cell(cA.row, 6).value
            n += 1 
   
def wordfinder_sum(searchString):
    n = 0
    for cA in list_colA:
        if cA.value == searchString:
            q = ws_a.cell(cA.row, 5).value
            p = ws_a.cell(cA.row, 6).value
            ws_f.cell(21 + n, 17).value = q * p
            n += 1

def wordfinder_total(searchString):
    sum_price = []
    for x in range(21,27):
        c = ws_f.cell(row=x, column=17).value
        if c != None:
            sum_price.append(c)
    ws_f['Q27'] = sum(sum_price)

def wordfinder_listprice(searchString):
    number = ws_f['Q27'].value
    number = str(number)
    strings = []
    n = 0
    for s in number:
        strings.append(s)
    strings.reverse()
    for w in strings:
        ws_f.cell(6, 15-n).value = w
        n += 1

def wordfinder_border():
    bian = Side(style='thin', color='000000') 
    bian1 = Side(style='medium', color='000000')
    border = Border(top=bian, bottom=bian, left=bian, right=bian, diagonal=bian, diagonalDown=True)
    border1 = Border(top=bian, bottom=bian1, left=bian, right=bian, diagonal=bian, diagonalDown=True)
    ws_f['B13'].border = border
    ws_f['B15'].border = border1
    ws_f['B33'].border = border
    ws_f['B34'].border = border1
    ws_f['E10'].border = border
    ws_f['E32'].border = border

def wordfinder_client(searchString):
    p = ws_f['Q27'].value
    for cA in list_colA:
        if cA.value == searchString:
            n1 = ws_a.cell(cA.row, 2).value
    ws_f['A29'].value = '2.    ■本案經詢價擬以' + str(p) + '元　交由　' + n1 + '辦理，並經驗收合格後付款。'
#產生黏貼憑證
for l in type_name:
    wb_t.save(filename = l + '黏貼憑證.xlsx')
    wb_f = load_workbook(filename = l + '黏貼憑證.xlsx', data_only=True)
    ws_f = wb_f['sheet']
    wordfinder_sort(l)
    wordfinder_name(l)
    wordfinder_type(l)
    wordfinder_quantity(l)
    wordfinder_price(l)
    wordfinder_sum(l)
    wordfinder_total(l)
    wordfinder_listprice(l)
    wordfinder_border()
    wordfinder_client(l)
    wb_f.save(filename = l + '黏貼憑證.xlsx')

#記帳簿
month = input('這是幾月份: ')
times = input('這是' + month +'月第幾次核銷? ')
month = int(month)
times = int(times)

#起始點
row1 = 0
for x in range(7,400):
    if ws.cell(row=x, column=2).value == None:
        row1 = x
        break 
t = 0
count = 0
def type_count(searchString):
    global t
    for cc in list_colA:
        if cc.value == searchString:
            t += 1


def type_in(searchString):
    global row1
    global count
    global t
    global month
    global times
    for a in range(2, 500):
        if searchString == ws_a.cell(a, 1).value and ws_a.cell(a, 7).value == '物品':
            ws.cell(row1, 2).value = month
            ws.cell(row1, 3).value = times
            ws.cell(row1, 4).value = ws_a.cell(a, 3).value
            ws.cell(row1, 8).value = ws_a.cell(a, 10).value
            row1 += 1
            count += 1
            if count > t :
                break
        elif searchString == ws_a.cell(a, 1).value and ws_a.cell(a, 7).value == '水費':
            ws.cell(row1, 2).value = month
            ws.cell(row1, 3).value = times
            ws.cell(row1, 4).value = ws_a.cell(a, 3).value
            ws.cell(row1, 5).value = ws_a.cell(a, 10).value
            row1 += 1
            count += 1
            if count > t :
                break
        elif searchString == ws_a.cell(a, 1).value and ws_a.cell(a, 7).value == '電費':
            ws.cell(row1, 2).value = month
            ws.cell(row1, 3).value = times
            ws.cell(row1, 4).value = ws_a.cell(a, 3).value
            ws.cell(row1, 6).value = ws_a.cell(a, 10).value
            row1 += 1
            count += 1
            if count > t :
                break
        elif searchString == ws_a.cell(a, 1).value and ws_a.cell(a, 7).value == '電話':
            ws.cell(row1, 2).value = month
            ws.cell(row1, 3).value = times
            ws.cell(row1, 4).value = ws_a.cell(a, 3).value
            ws.cell(row1, 7).value = ws_a.cell(a, 10).value
            row1 += 1
            count += 1
            if count > t :
                break
        elif searchString == ws_a.cell(a, 1).value and ws_a.cell(a, 7).value == '一般':
            ws.cell(row1, 2).value = month
            ws.cell(row1, 3).value = times
            ws.cell(row1, 4).value = ws_a.cell(a, 3).value
            ws.cell(row1, 9).value = ws_a.cell(a, 10).value
            row1 += 1
            count += 1
            if count > t :
                break

def border_type():
    for x in range(7,500):
        for y in range(1,11):
            ws.cell(row=x, column=y).border = border3

def type_times():
    global month
    global times
    ws2 = wb['大隊明細表']
    c = 0
    for aa in range(10, 30):
        if c > 0:
            break
        elif ws2.cell(2,aa).value == None and c == 0 :
            ws2.cell(2,aa).value = 1
            ws2.cell(2,aa-1).value = 0
            ws2.cell(5,aa).value = month
            ws2.cell(6,aa).value = times
            c += 1

for a in type_name:
    type_count(a)
    type_in(a)
    border_type()
print('本次填入', t, '個資料。 ' )
type_times()

wb.save(str(month) + '月' + str(times) + '次.xlsx')
source = os.getcwd()
#source = unicode(source, 'utf-8') #路徑有亂碼時加入本行
print(source)
files = os.listdir(source)       
for file in files: # '.'代表現在資料夾的位子
    if file.endswith('黏貼憑證.xlsx'):
        shutil.move(file, './黏貼憑證')
    elif file.endswith('次.xlsx'):
        shutil.move(file, './日記帳') 



        

